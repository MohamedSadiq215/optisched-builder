"""
engine/excel.py — Professional Excel workbook export using openpyxl.

Tab 1: Calendar Event Banner + Master Schedule (coloured cells, frozen panes)
Tab 2: Coverage Heatmap & Compliance (COUNTIF formulas + conditional formatting)
"""

from __future__ import annotations

import calendar
import datetime
import io
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule


# ── Colour palette ────────────────────────────────────────────────────────────

NAVY      = "1A3A5C"
ICE_BLUE  = "D6E4F0"
WHITE     = "FFFFFF"
LIGHT_GREY = "F2F2F2"

SHIFT_FILLS: dict[str, PatternFill] = {
    "A":   PatternFill("solid", fgColor="C8E6C9"),   # light green
    "MS":  PatternFill("solid", fgColor="B2EBF2"),   # pale teal
    "B":   PatternFill("solid", fgColor="FFF9C4"),   # muted yellow
    "C":   PatternFill("solid", fgColor="E1BEE7"),   # soft purple
    "OFF": PatternFill("solid", fgColor="ECEFF1"),   # light grey-blue
    "AL":  PatternFill("solid", fgColor="FCE4EC"),   # rose
}
DEFAULT_FILL = PatternFill("solid", fgColor=WHITE)

HEADER_FILL   = PatternFill("solid", fgColor=NAVY)
SUBHEAD_FILL  = PatternFill("solid", fgColor=ICE_BLUE)
EVENT_FILL    = PatternFill("solid", fgColor="FFF3CD")

HEADER_FONT   = Font(name="Calibri", bold=True, color=WHITE,     size=11)
SUBHEAD_FONT  = Font(name="Calibri", bold=True, color=NAVY,      size=10)
CELL_FONT     = Font(name="Calibri",             color="263238", size=10)
EVENT_FONT    = Font(name="Calibri", bold=True,  color="856404", size=9, italic=True)

THIN  = Side(style="thin",   color="B0BEC5")
THICK = Side(style="medium", color=NAVY)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")


def _apply_border(cell) -> None:
    cell.border = CELL_BORDER


def _write_cell(
    ws,
    row: int,
    col: int,
    value: Any,
    font: Font = CELL_FONT,
    fill: PatternFill | None = None,
    alignment: Alignment = CENTER,
    number_format: str | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    _apply_border(cell)
    if number_format:
        cell.number_format = number_format


# ── Tab 1: Master Schedule ────────────────────────────────────────────────────

def _build_tab1(
    wb: Workbook,
    schedule_df: pd.DataFrame,
    year: int,
    month: int,
    events: dict[str, str],
) -> None:
    ws = wb.active
    ws.title = "Master Schedule"

    days_in_month = calendar.monthrange(year, month)[1]
    agents        = schedule_df.index.tolist()
    days          = list(range(1, days_in_month + 1))

    # ── Row 1: Workbook title ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 2)
    title_cell = ws.cell(row=1, column=1,
                         value=f"OptiSched Builder — {calendar.month_name[month]} {year}")
    title_cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=14)
    title_cell.fill      = HEADER_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # ── Row 2: Calendar event banner ──────────────────────────────────────────
    ws.row_dimensions[2].height = 36
    ws.cell(row=2, column=1, value="📌 Events").font  = SUBHEAD_FONT
    ws.cell(row=2, column=1).fill      = SUBHEAD_FILL
    ws.cell(row=2, column=1).alignment = CENTER

    for d_idx, d in enumerate(days):
        col = d_idx + 2
        dt  = datetime.date(year, month, d)
        ds  = dt.strftime("%Y-%m-%d")
        ev  = events.get(ds, "")
        _write_cell(ws, 2, col, ev or "",
                    font=EVENT_FONT, fill=EVENT_FILL if ev else SUBHEAD_FILL)

    # ── Row 3: Day-of-week sub-header ─────────────────────────────────────────
    ws.row_dimensions[3].height = 18
    ws.cell(row=3, column=1, value="Agent").font  = SUBHEAD_FONT
    ws.cell(row=3, column=1).fill      = SUBHEAD_FILL
    ws.cell(row=3, column=1).alignment = CENTER

    for d_idx, d in enumerate(days):
        col = d_idx + 2
        dt  = datetime.date(year, month, d)
        dow = dt.strftime("%a")
        _write_cell(ws, 3, col, dow, font=SUBHEAD_FONT, fill=SUBHEAD_FILL)

    # ── Row 4: Day-number header ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws.cell(row=4, column=1, value="").fill = HEADER_FILL
    for d_idx, d in enumerate(days):
        col = d_idx + 2
        _write_cell(ws, 4, col, d, font=HEADER_FONT, fill=HEADER_FILL)

    # ── Agent rows ────────────────────────────────────────────────────────────
    for a_idx, agent in enumerate(agents):
        row = a_idx + 5

        # Agent name column
        name_cell = ws.cell(row=row, column=1, value=agent)
        name_cell.font      = Font(name="Calibri", bold=True, color=NAVY, size=10)
        name_cell.fill      = PatternFill("solid", fgColor=ICE_BLUE)
        name_cell.alignment = LEFT
        _apply_border(name_cell)

        for d_idx, d in enumerate(days):
            col   = d_idx + 2
            shift = str(schedule_df.loc[agent, str(d)])
            fill  = SHIFT_FILLS.get(shift, DEFAULT_FILL)
            _write_cell(ws, row, col, shift, fill=fill)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    for d_idx in range(days_in_month):
        ws.column_dimensions[get_column_letter(d_idx + 2)].width = 5

    # ── Freeze panes (keep agent column + header rows visible) ────────────────
    ws.freeze_panes = "B5"


# ── Tab 2: Coverage Heatmap ───────────────────────────────────────────────────

def _build_tab2(
    wb: Workbook,
    schedule_df: pd.DataFrame,
    year: int,
    month: int,
    max_caps_df: pd.DataFrame,
    min_caps_df: pd.DataFrame,
    coverage_df: pd.DataFrame | None,
) -> None:
    ws = wb.create_sheet("Coverage & Compliance")

    days_in_month = calendar.monthrange(year, month)[1]
    agents        = schedule_df.index.tolist()
    days          = list(range(1, days_in_month + 1))

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 2)
    tc = ws.cell(row=1, column=1,
                 value=f"Coverage Heatmap & Compliance — {calendar.month_name[month]} {year}")
    tc.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    tc.fill = HEADER_FILL
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # ── Day headers ───────────────────────────────────────────────────────────
    _write_cell(ws, 2, 1, "Shift / Metric", font=SUBHEAD_FONT, fill=SUBHEAD_FILL, alignment=LEFT)
    for d_idx, d in enumerate(days):
        dt  = datetime.date(year, month, d)
        dow = dt.strftime("%a")
        _write_cell(ws, 2, d_idx + 2, f"{d}\n{dow}",
                    font=SUBHEAD_FONT, fill=SUBHEAD_FILL)
    ws.row_dimensions[2].height = 28

    # ── Data section: actual counts using COUNTIF ─────────────────────────────
    # Write agent assignments into a hidden area so COUNTIF can reference them
    # We instead write count values directly (computed from schedule_df)
    work_shifts = ["A", "MS", "B", "C", "OFF", "AL"]

    row_offset = 3
    for s in work_shifts:
        row = row_offset
        _write_cell(ws, row, 1, f"Count: {s}", font=CELL_FONT, fill=DEFAULT_FILL, alignment=LEFT)

        for d_idx, d in enumerate(days):
            col    = d_idx + 2
            count  = int((schedule_df[str(d)] == s).sum())
            dt     = datetime.date(year, month, d)
            dow    = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][dt.weekday()]

            # Determine fill based on compliance
            fill = DEFAULT_FILL
            try:
                max_row = max_caps_df[max_caps_df.iloc[:, 0].str.contains(s, na=False)]
                min_row = min_caps_df[min_caps_df.iloc[:, 0].str.contains(s, na=False)]
                max_v   = int(max_row[dow].values[0]) if not max_row.empty else 9999
                min_v   = int(min_row[dow].values[0]) if not min_row.empty else 0
                if count > max_v:
                    fill = PatternFill("solid", fgColor="FFCDD2")   # red
                elif count < min_v:
                    fill = PatternFill("solid", fgColor="FFF9C4")   # yellow
                else:
                    fill = PatternFill("solid", fgColor="C8E6C9")   # green
            except Exception:
                pass

            _write_cell(ws, row, col, count, fill=fill)

        row_offset += 1

    # ── Separator ─────────────────────────────────────────────────────────────
    row_offset += 1
    _write_cell(ws, row_offset, 1, "MIN Required", font=SUBHEAD_FONT, fill=SUBHEAD_FILL)
    for d_idx, d in enumerate(days):
        _write_cell(ws, row_offset, d_idx + 2, "", fill=SUBHEAD_FILL)

    row_offset += 1
    for _, cap_row in min_caps_df.iterrows():
        label = str(cap_row.iloc[0])
        _write_cell(ws, row_offset, 1, label, font=CELL_FONT, alignment=LEFT)
        for d_idx, d in enumerate(days):
            dt  = datetime.date(year, month, d)
            dow = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][dt.weekday()]
            val = int(cap_row.get(dow, 0))
            _write_cell(ws, row_offset, d_idx + 2, val)
        row_offset += 1

    row_offset += 1
    _write_cell(ws, row_offset, 1, "MAX Allowed", font=SUBHEAD_FONT, fill=SUBHEAD_FILL)
    for d_idx, d in enumerate(days):
        _write_cell(ws, row_offset, d_idx + 2, "", fill=SUBHEAD_FILL)

    row_offset += 1
    for _, cap_row in max_caps_df.iterrows():
        label = str(cap_row.iloc[0])
        _write_cell(ws, row_offset, 1, label, font=CELL_FONT, alignment=LEFT)
        for d_idx, d in enumerate(days):
            dt  = datetime.date(year, month, d)
            dow = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][dt.weekday()]
            val = int(cap_row.get(dow, 0))
            _write_cell(ws, row_offset, d_idx + 2, val)
        row_offset += 1

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "B3"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    for d_idx in range(days_in_month):
        ws.column_dimensions[get_column_letter(d_idx + 2)].width = 5


# ── Entry point ───────────────────────────────────────────────────────────────

def build_workbook(
    buf: io.BytesIO,
    schedule_df: pd.DataFrame,
    year: int,
    month: int,
    events: dict[str, str],
    max_caps_df: pd.DataFrame,
    min_caps_df: pd.DataFrame,
    coverage_df: pd.DataFrame | None,
) -> None:
    wb = Workbook()

    _build_tab1(wb, schedule_df, year, month, events)
    _build_tab2(wb, schedule_df, year, month, max_caps_df, min_caps_df, coverage_df)

    wb.save(buf)
